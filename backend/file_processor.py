"""
File Upload & Document Processor — SwarmOps
Supports: PDF, CSV, DOCX, XLSX, TXT, JSON, PNG/JPG
Extracted content is injected into agent prompts automatically.
"""
import os
import io
import csv
import json
import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# Upload directory (relative to project root when running via Railway/Render)
_BASE = Path(__file__).parent.parent  # project root
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", str(_BASE / "uploads")))
try:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    UPLOAD_DIR = Path("uploads")
    UPLOAD_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {'.pdf', '.csv', '.txt', '.docx', '.xlsx', '.json', '.png', '.jpg', '.jpeg'}
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_CONTENT_CHARS   = 5_000             # chars injected into prompt
MAX_PAGES_PDF       = 30


class FileProcessor:
    """Process uploaded files and make content available to agents."""

    def __init__(self):
        self.uploaded_files: list[dict] = []

    # ── Main entry point ───────────────────────────────────────────────────

    async def process_upload(self, file) -> dict:
        """
        Process a FastAPI UploadFile.
        Returns a result dict with success, content, summary, metadata.
        """
        filename = getattr(file, "filename", None) or "unknown"
        ext = Path(filename).suffix.lower()

        if ext not in ALLOWED_EXTENSIONS:
            return {
                "success": False,
                "error": f"'{ext}' not supported. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
            }

        try:
            raw = await file.read()
        except Exception as e:
            return {"success": False, "error": f"Read error: {e}"}

        if len(raw) > MAX_FILE_SIZE_BYTES:
            mb = len(raw) / 1024 / 1024
            return {"success": False, "error": f"File too large ({mb:.1f} MB). Max 10 MB."}

        # Persist to disk
        safe_name = datetime.utcnow().strftime("%Y%m%d_%H%M%S_") + Path(filename).name
        file_path = UPLOAD_DIR / safe_name
        try:
            file_path.write_bytes(raw)
        except Exception as e:
            logger.warning(f"[file_processor] Could not save to disk: {e}")

        # Extract text
        extracted, metadata = self._extract(raw, ext)

        word_count = len(extracted.split()) if extracted else 0

        result = {
            "success":              True,
            "filename":             filename,
            "file_type":            ext,
            "file_size":            len(raw),
            "content":              extracted[:MAX_CONTENT_CHARS],
            "full_content_length":  len(extracted),
            "word_count":           word_count,
            "metadata":             metadata,
            "summary":              self._summary(extracted, filename, ext),
            "uploaded_at":          datetime.utcnow().isoformat(),
        }
        self.uploaded_files.append(result)
        logger.info(f"[file_processor] Processed {filename} ({word_count} words, {ext})")
        return result

    # ── Extraction dispatch ────────────────────────────────────────────────

    def _extract(self, raw: bytes, ext: str) -> tuple[str, dict]:
        dispatch = {
            ".txt":  lambda r: (self._txt(r), {}),
            ".csv":  lambda r: self._csv(r),
            ".json": lambda r: self._json(r),
            ".pdf":  lambda r: (self._pdf(r), {}),
            ".docx": lambda r: (self._docx(r), {}),
            ".xlsx": lambda r: self._xlsx(r),
        }
        if ext in dispatch:
            try:
                return dispatch[ext](raw)
            except Exception as e:
                return f"[Extraction error: {e}]", {}
        # Images
        return "[Image uploaded. Describe its contents to include in analysis.]", {}

    def _txt(self, raw: bytes) -> str:
        for enc in ("utf-8", "latin-1", "cp1252"):
            try:
                return raw.decode(enc)
            except UnicodeDecodeError:
                continue
        return "[Could not decode text file]"

    def _csv(self, raw: bytes) -> tuple[str, dict]:
        try:
            text = raw.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return "Empty CSV.", {}
            headers = list(rows[0].keys())
            meta = {"headers": headers, "row_count": len(rows), "column_count": len(headers)}
            lines = [
                f"CSV: {len(rows)} rows × {len(headers)} columns.",
                f"Columns: {', '.join(headers)}",
                "",
            ]
            for i, row in enumerate(rows[:25]):
                cells = " | ".join(f"{k}: {v}" for k, v in row.items() if v not in (None, ""))
                lines.append(f"Row {i+1}: {cells}")
            if len(rows) > 25:
                lines.append(f"… and {len(rows)-25} more rows")
            return "\n".join(lines), meta
        except Exception as e:
            return f"[CSV error: {e}]", {}

    def _json(self, raw: bytes) -> tuple[str, dict]:
        try:
            data = json.loads(raw.decode("utf-8", errors="replace"))
            if isinstance(data, list):
                meta = {"type": "array", "length": len(data)}
                return f"JSON array ({len(data)} items):\n" + json.dumps(data[:15], indent=2)[:3000], meta
            elif isinstance(data, dict):
                keys = list(data.keys())
                meta = {"type": "object", "keys": keys[:20]}
                return f"JSON object — keys: {', '.join(keys[:20])}\n\n" + json.dumps(data, indent=2)[:3000], meta
            return str(data)[:3000], {"type": type(data).__name__}
        except Exception as e:
            return f"[JSON error: {e}]", {}

    def _pdf(self, raw: bytes) -> str:
        # Try pdfplumber first, PyPDF2 as fallback
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                parts = []
                for i, page in enumerate(pdf.pages[:MAX_PAGES_PDF]):
                    t = page.extract_text() or ""
                    if t.strip():
                        parts.append(f"[Page {i+1}]\n{t.strip()}")
                return "\n\n".join(parts) or "[PDF: no extractable text]"
        except ImportError:
            pass
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(raw))
            parts = []
            for i, page in enumerate(reader.pages[:MAX_PAGES_PDF]):
                t = page.extract_text() or ""
                if t.strip():
                    parts.append(f"[Page {i+1}]\n{t.strip()}")
            return "\n\n".join(parts) or "[PDF: no extractable text]"
        except ImportError:
            return "[PDF requires pdfplumber or PyPDF2 — pip install pdfplumber]"
        except Exception as e:
            return f"[PDF error: {e}]"

    def _docx(self, raw: bytes) -> str:
        try:
            from docx import Document
            doc = Document(io.BytesIO(raw))
            paras = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paras) or "[DOCX: no text]"
        except ImportError:
            return "[DOCX requires python-docx — pip install python-docx]"
        except Exception as e:
            return f"[DOCX error: {e}]"

    def _xlsx(self, raw: bytes) -> tuple[str, dict]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            all_text = []
            meta: dict = {"sheets": []}
            for sheet_name in wb.sheetnames[:5]:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True, max_row=50))
                if not rows:
                    continue
                headers = [str(c) if c is not None else "" for c in rows[0]]
                meta["sheets"].append({"name": sheet_name, "columns": headers})
                all_text.append(f"\n=== Sheet: {sheet_name} ===")
                all_text.append(f"Columns: {', '.join(h for h in headers if h)}")
                for i, row in enumerate(rows[1:21]):
                    cells = " | ".join(
                        f"{headers[j]}: {v}"
                        for j, v in enumerate(row)
                        if v is not None and j < len(headers) and headers[j]
                    )
                    if cells:
                        all_text.append(f"  Row {i+1}: {cells}")
            wb.close()
            return "\n".join(all_text), meta
        except ImportError:
            return "[Excel requires openpyxl — pip install openpyxl]", {}
        except Exception as e:
            return f"[Excel error: {e}]", {}

    # ── Summary ────────────────────────────────────────────────────────────

    def _summary(self, text: str, filename: str, ext: str) -> str:
        wc = len(text.split())
        msgs = {
            ".csv":  f"CSV file '{filename}' — use this data for quantitative analysis.",
            ".pdf":  f"PDF document '{filename}' ({wc:,} words) — analyze for marketing insights.",
            ".docx": f"Word document '{filename}' ({wc:,} words) — review and incorporate.",
            ".xlsx": f"Excel spreadsheet '{filename}' — use for data-driven recommendations.",
            ".json": f"JSON data file '{filename}' — use this structured data in analysis.",
        }
        if ext in {".png", ".jpg", ".jpeg"}:
            return f"Image '{filename}' uploaded for reference."
        return msgs.get(ext, f"File '{filename}' ({wc:,} words) — use as analysis context.")

    # ── Context builder ────────────────────────────────────────────────────

    def get_context_block(self) -> str:
        """Return all uploaded files as an agent-ready context block."""
        ok = [f for f in self.uploaded_files if f.get("success")]
        if not ok:
            return ""
        lines = ["=== UPLOADED DOCUMENTS (use this data in your analysis) ==="]
        for f in ok:
            lines.append(f"\n--- {f['filename']} ({f['file_type']}, {f['word_count']:,} words) ---")
            lines.append(f["content"][:MAX_CONTENT_CHARS])
        lines.append("=== END UPLOADED DOCUMENTS ===")
        return "\n".join(lines)

    @property
    def count(self) -> int:
        return sum(1 for f in self.uploaded_files if f.get("success"))


# ── Singleton ──────────────────────────────────────────────────────────────
_instance: FileProcessor | None = None

def get_file_processor() -> FileProcessor:
    global _instance
    if _instance is None:
        _instance = FileProcessor()
    return _instance

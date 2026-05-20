"""Universal file upload processor."""
import io
import csv
import json
import logging
from pathlib import Path
from typing import Dict

logger = logging.getLogger(__name__)

ALLOWED = {
    '.pdf', '.csv', '.txt', '.docx', '.xlsx', '.json', '.md', '.html',
    '.py', '.js', '.ts', '.sql', '.yaml', '.yml',
    '.png', '.jpg', '.jpeg', '.gif', '.webp',
}
MAX_SIZE = 25 * 1024 * 1024  # 25MB


async def process_file(file) -> Dict:
    """Extract text content from uploaded file."""
    filename = file.filename or "unknown"
    ext = Path(filename).suffix.lower()

    if ext not in ALLOWED:
        return {"success": False, "error": f"File type {ext} not supported"}

    content_bytes = await file.read()
    if len(content_bytes) > MAX_SIZE:
        size_mb = len(content_bytes) / 1024 / 1024
        return {"success": False, "error": f"File too large ({size_mb:.1f}MB, max 25MB)"}

    extracted = ""
    metadata = {}

    try:
        if ext in {'.txt', '.md', '.html', '.py', '.js', '.ts', '.sql', '.yaml', '.yml'}:
            extracted = _extract_text(content_bytes)
        elif ext == '.csv':
            extracted, metadata = _extract_csv(content_bytes)
        elif ext == '.json':
            extracted, metadata = _extract_json(content_bytes)
        elif ext == '.pdf':
            extracted = _extract_pdf(content_bytes)
        elif ext == '.docx':
            extracted = _extract_docx(content_bytes)
        elif ext == '.xlsx':
            extracted, metadata = _extract_xlsx(content_bytes)
        elif ext in {'.png', '.jpg', '.jpeg', '.gif', '.webp'}:
            extracted = f"[Image: {filename}. Image analysis not available in v1.]"

    except Exception as e:
        logger.error(f"Extraction failed for {filename}: {e}")
        return {"success": False, "error": str(e)}

    return {
        "success": True,
        "filename": filename,
        "file_type": ext,
        "file_size": len(content_bytes),
        "content": extracted[:5000],
        "full_length": len(extracted),
        "word_count": len(extracted.split()),
        "metadata": metadata,
        "summary": f"{ext.upper().strip('.')} file '{filename}' ({len(extracted.split())} words extracted)",
    }


def _extract_text(content: bytes) -> str:
    try:
        return content.decode('utf-8')
    except UnicodeDecodeError:
        return content.decode('latin-1', errors='replace')


def _extract_csv(content: bytes) -> tuple:
    text = content.decode('utf-8', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        return "Empty CSV", {}

    headers = list(rows[0].keys())
    metadata = {"headers": headers, "row_count": len(rows)}

    lines = [f"CSV with {len(rows)} rows, columns: {', '.join(headers)}", ""]
    for i, row in enumerate(rows[:25]):
        lines.append(f"Row {i+1}: " + " | ".join(f"{k}={v}" for k, v in row.items()))
    if len(rows) > 25:
        lines.append(f"...and {len(rows)-25} more rows")

    return "\n".join(lines), metadata


def _extract_json(content: bytes) -> tuple:
    try:
        data = json.loads(content.decode('utf-8'))
        if isinstance(data, list):
            metadata = {"type": "array", "items": len(data)}
            text = f"JSON array, {len(data)} items:\n\n{json.dumps(data[:10], indent=2)[:3000]}"
        elif isinstance(data, dict):
            metadata = {"type": "object", "keys": list(data.keys())[:20]}
            text = json.dumps(data, indent=2)[:3000]
        else:
            metadata = {"type": type(data).__name__}
            text = str(data)[:3000]
        return text, metadata
    except Exception as e:
        return f"[JSON error: {e}]", {}


def _extract_pdf(content: bytes) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            texts = []
            for i, page in enumerate(pdf.pages[:30]):
                t = page.extract_text() or ""
                if t:
                    texts.append(f"--- Page {i+1} ---\n{t}")
            return "\n\n".join(texts) if texts else "[Empty PDF]"
    except ImportError:
        return "[PDF support requires pdfplumber]"


def _extract_docx(content: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(content))
        paras = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paras) if paras else "[Empty DOCX]"
    except ImportError:
        return "[DOCX support requires python-docx]"


def _extract_xlsx(content: bytes) -> tuple:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        all_text = []
        metadata = {"sheets": []}

        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True, max_row=50))
            if not rows:
                continue
            metadata["sheets"].append({"name": sheet_name, "rows": ws.max_row})
            all_text.append(f"\n=== Sheet: {sheet_name} ===")
            for i, row in enumerate(rows[:20]):
                all_text.append(f"Row {i+1}: " + " | ".join(str(v) for v in row if v is not None))

        wb.close()
        return "\n".join(all_text), metadata
    except ImportError:
        return "[Excel support requires openpyxl]", {}
